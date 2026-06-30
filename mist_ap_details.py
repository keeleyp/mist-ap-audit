#!/usr/bin/env python3
"""
mist_ap_details.py — Juniper Mist AP Details Report
=====================================================
Queries the Mist API site-by-site and produces a single-sheet Excel workbook
containing detailed AP inventory and health data for every AP in the organisation.

Unlike mist_ap_report.py (which uses the org-level stats endpoint), this script
calls the per-site stats endpoint so it can capture fields that are only returned
at site scope — port stats, LLDP neighbour info, PoE draw, ESL state, etc.

Pre-flight:
  - Fetches org name and site count
  - Estimates total API calls required
  - Checks calls remaining in the current rate-limit window (/self/usage)
  - Blocks with a wait/quit prompt if there are insufficient calls available

During the run:
  - Re-checks the rate limit every 50 sites
  - If calls run out mid-run, offers to wait for the window to reset then continue

Output:
  Mist_AP_Details_<OrgName>_<YYYY-MM-DD_HHMMSS>.xlsx

Config file: mist_ap_report.ini (shared with mist_ap_report.py)
  [mist]   api_base, org_id, api_token
  [output] directory
  (The [thresholds] section is not used by this script.)

Usage:
  python3 mist_ap_details.py

Requirements:
  pip install requests openpyxl
"""
import configparser
import requests
import time
import os
import sys
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

config = configparser.ConfigParser()
config.read(os.path.join(os.path.dirname(os.path.abspath(__file__)), "mist_ap_report.ini"))

API_BASE   = config.get("mist", "api_base")
ORG_ID     = config.get("mist", "org_id")
API_TOKEN  = config.get("mist", "api_token")
OUTPUT_DIR = os.path.expanduser(config.get("output", "directory"))

HEADERS = {"Authorization": f"Token {API_TOKEN}"}

HEADER_COLOR  = "1A237E"
RATE_HEADROOM = 10   # keep this many calls in reserve before pausing

# Issue-flag fills
FILL_OK      = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")   # green  ✓
FILL_ISSUE   = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")   # red    ✗
FILL_FLAGGED = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")   # amber  (problem cell)
FONT_WHITE   = Font(bold=True, color="FFFFFF", size=12)

# Columns to check and the condition that marks them as a problem.
# A condition returns True when the value indicates an issue.
# Empty/missing values are skipped (AP may be offline or have no eth0 data).
ISSUE_CHECKS = {
    "eth0 Full Duplex":  lambda v: v is not True,
    "Power Constrained": lambda v: v is True,
    "eth0 Speed (Mbps)": lambda v: v == 100,
    "eth0 RX Errors":    lambda v: isinstance(v, (int, float)) and v != 0,
}

def flagged_columns(row):
    """Return the set of column names that have a health issue for this AP row."""
    flags = set()
    for col, check in ISSUE_CHECKS.items():
        val = row.get(col)
        if val == "" or val is None:
            continue   # no data — skip rather than false-flag
        if check(val):
            flags.add(col)
    return flags

# ---------------------------------------------------------------------------
# Rate-limit helpers
# ---------------------------------------------------------------------------

def fetch_api_usage():
    """Return (used, limit, secs_elapsed) from /self/usage."""
    resp = requests.get(f"{API_BASE}/self/usage", headers=HEADERS)
    resp.raise_for_status()
    d = resp.json()
    return d.get("requests", 0), d.get("request_limit", 5000), d.get("seconds", 0)

def remaining_and_reset(used, limit, _secs_elapsed=None):
    remaining = limit - used
    # Reset happens on the clock hour, not 60 mins from first request.
    now = datetime.now()
    secs_into_hour = now.minute * 60 + now.second
    secs_left = max(0, 3600 - secs_into_hour)
    return remaining, secs_left

def wait_for_reset(secs_left, reason=""):
    """Block until the rate-limit window resets, showing a live countdown."""
    wait = secs_left + 15   # 15-second buffer after reset
    if reason:
        print(f"\n  {reason}")
    print(f"  Waiting {wait // 60}m {wait % 60}s for the rate-limit window to reset...")
    deadline = time.time() + wait
    try:
        while True:
            left = int(deadline - time.time())
            if left <= 0:
                break
            m, s = divmod(left, 60)
            sys.stdout.write(f"\r  Resuming in {m}m {s:02d}s ...   ")
            sys.stdout.flush()
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n\n  Wait interrupted by user. Exiting.")
        sys.exit(1)
    print("\r  Rate-limit window has reset. Continuing...          ")

def check_rate_limit(calls_needed, context=""):
    """
    Fetch current usage and block if insufficient calls remain.
    Returns the refreshed (remaining, secs_left).
    Exits if the user chooses to quit.
    """
    used, limit, secs_elapsed = fetch_api_usage()
    remaining, secs_left = remaining_and_reset(used, limit, secs_elapsed)

    if remaining - RATE_HEADROOM >= calls_needed:
        return remaining, secs_left   # enough headroom — carry on

    m, s = divmod(secs_left, 60)
    label = f" ({context})" if context else ""
    print(f"\n\n  {'='*56}")
    print(f"  *** RATE LIMIT WARNING{label} ***")
    print(f"  Calls needed:    {calls_needed}")
    print(f"  Calls remaining: {remaining}  (limit {limit}/hour, used {used})")
    print(f"  Window resets in ~{m}m {s}s")
    print(f"  {'='*56}")
    print("  Options:")
    print("    w — wait for the window to reset then continue automatically")
    print("    q — quit now (no report will be saved)")

    while True:
        choice = input("  Your choice (w/q): ").strip().lower()
        if choice == "q":
            print("  Aborted by user.")
            sys.exit(0)
        if choice == "w":
            wait_for_reset(secs_left)
            # refresh after waiting
            used, limit, secs_elapsed = fetch_api_usage()
            remaining, secs_left = remaining_and_reset(used, limit, secs_elapsed)
            print(f"  New remaining calls: {remaining}")
            return remaining, secs_left
        print("  Please enter 'w' to wait or 'q' to quit.")

# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def fetch_org_info():
    resp = requests.get(f"{API_BASE}/orgs/{ORG_ID}", headers=HEADERS)
    resp.raise_for_status()
    return resp.json()

def fetch_org_stats():
    resp = requests.get(f"{API_BASE}/orgs/{ORG_ID}/stats", headers=HEADERS)
    resp.raise_for_status()
    return resp.json()

def fetch_all_sites():
    """Return list of {id, name} dicts for all sites in the org."""
    sites = []
    page  = 1
    while True:
        url  = f"{API_BASE}/orgs/{ORG_ID}/sites?limit=1000&page={page}"
        resp = requests.get(url, headers=HEADERS)
        resp.raise_for_status()
        data = resp.json()
        if not data:
            break
        sites.extend(data)
        sys.stdout.write(f"\r  Fetched {len(sites)} sites (page {page})   ")
        sys.stdout.flush()
        if len(data) < 1000:
            break
        page += 1
    print(f"\r  {len(sites)} sites fetched.                        ")
    return sites

def fetch_site_ap_stats(site_id, site_name, idx, total):
    """Fetch all AP stats for a single site (≤1000 APs assumed)."""
    url  = f"{API_BASE}/sites/{site_id}/stats/devices?type=ap&limit=1000"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    data = resp.json()
    sys.stdout.write(f"\r  [{idx}/{total}] {site_name[:40]:<40} — {len(data)} APs   ")
    sys.stdout.flush()
    return data

def format_eta(seconds):
    if seconds < 60:
        return f"{int(seconds)}s"
    m, s = divmod(int(seconds), 60)
    return f"{m}m {s}s"

# ---------------------------------------------------------------------------
# Data extraction helpers
# ---------------------------------------------------------------------------

def extract_eth0_port(ap):
    eth0 = ap.get("port_stat", {}).get("eth0", {})
    if not eth0:
        return {}
    return {
        "eth0_up":         eth0.get("up", ""),
        "eth0_speed":      eth0.get("speed", ""),
        "eth0_full_duplex":eth0.get("full_duplex", ""),
        "eth0_tx_bytes":   eth0.get("tx_bytes", ""),
        "eth0_tx_pkts":    eth0.get("tx_pkts", ""),
        "eth0_rx_bytes":   eth0.get("rx_bytes", ""),
        "eth0_rx_pkts":    eth0.get("rx_pkts", ""),
        "eth0_rx_errors":  eth0.get("rx_errors", ""),
        "eth0_rx_peak_bps":eth0.get("rx_peak_bps", ""),
        "eth0_tx_peak_bps":eth0.get("tx_peak_bps", ""),
    }

def extract_lldp(ap):
    lldp = ap.get("lldp_stats", {}).get("eth0", {})
    if not lldp:
        lldp = ap.get("lldp_stat", {})
    if not lldp:
        return {}
    return {
        "lldp_system_name":       lldp.get("system_name", ""),
        "lldp_system_desc":       lldp.get("system_desc", ""),
        "lldp_mgmt_addr":         lldp.get("mgmt_addr", ""),
        "lldp_port_desc":         lldp.get("port_desc", ""),
        "lldp_port_id":           lldp.get("port_id", ""),
        "lldp_chassis_id":        lldp.get("chassis_id", ""),
        "lldp_med_supported":     lldp.get("lldp_med_supported", ""),
        "lldp_power_req_count":   lldp.get("power_request_count", ""),
        "lldp_power_allocated_mw":lldp.get("power_allocated", ""),
        "lldp_power_requested_mw":lldp.get("power_requested", ""),
        "lldp_power_draw_mw":     lldp.get("power_draw", ""),
    }

def extract_ip_stat(ap):
    ip_stat = ap.get("ip_stat", {})
    dns = ip_stat.get("dns", [])
    return {
        "ip_gateway":  ip_stat.get("gateway", ""),
        "ip_netmask":  ip_stat.get("netmask", ""),
        "ip_dns":      ", ".join(dns) if dns else "",
        "ip_dhcp_srv": ip_stat.get("dhcp_server", ""),
    }

def flatten_ap(ap, site_name):
    ip_stat = extract_ip_stat(ap)
    eth0    = extract_eth0_port(ap)
    lldp    = extract_lldp(ap)

    inactive_vlans = ap.get("inactive_wired_vlans", [])
    inactive_str   = ", ".join(str(v) for v in inactive_vlans) if inactive_vlans else ""

    esl     = ap.get("esl_stat", {})
    esl_str = json.dumps(esl) if esl else ""

    return {
        "Site Name":             site_name,
        "AP Name":               ap.get("name", ""),
        "MAC":                   ap.get("mac", ""),
        "Status":                ap.get("status", ""),
        "Power Constrained":     ap.get("power_constrained", ""),
        "IP":                    ap.get("ip", ""),
        "Gateway":               ip_stat.get("ip_gateway", ""),
        "Netmask":               ip_stat.get("ip_netmask", ""),
        "DNS Servers":           ip_stat.get("ip_dns", ""),
        "DHCP Server":           ip_stat.get("ip_dhcp_srv", ""),
        "External IP":           ap.get("ext_ip", ""),
        "Mount":                 ap.get("mount", ""),
        "eth0 Up":               eth0.get("eth0_up", ""),
        "eth0 Speed (Mbps)":     eth0.get("eth0_speed", ""),
        "eth0 Full Duplex":      eth0.get("eth0_full_duplex", ""),
        "eth0 TX Bytes":         eth0.get("eth0_tx_bytes", ""),
        "eth0 TX Pkts":          eth0.get("eth0_tx_pkts", ""),
        "eth0 RX Bytes":         eth0.get("eth0_rx_bytes", ""),
        "eth0 RX Pkts":          eth0.get("eth0_rx_pkts", ""),
        "eth0 RX Errors":        eth0.get("eth0_rx_errors", ""),
        "eth0 RX Peak (bps)":    eth0.get("eth0_rx_peak_bps", ""),
        "eth0 TX Peak (bps)":    eth0.get("eth0_tx_peak_bps", ""),
        "Inactive Wired VLANs":  inactive_str,
        "Switch Name":           lldp.get("lldp_system_name", ""),
        "Switch Description":    lldp.get("lldp_system_desc", ""),
        "Switch Mgmt Addr":      lldp.get("lldp_mgmt_addr", ""),
        "Switch Port Desc":      lldp.get("lldp_port_desc", ""),
        "Switch Port ID":        lldp.get("lldp_port_id", ""),
        "Switch Chassis ID":     lldp.get("lldp_chassis_id", ""),
        "LLDP MED Supported":    lldp.get("lldp_med_supported", ""),
        "PoE Req Count":         lldp.get("lldp_power_req_count", ""),
        "PoE Allocated (mW)":    lldp.get("lldp_power_allocated_mw", ""),
        "PoE Requested (mW)":    lldp.get("lldp_power_requested_mw", ""),
        "PoE Draw (mW)":         lldp.get("lldp_power_draw_mw", ""),
        "ESL Stat":              esl_str,
    }

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

COLUMNS = [
    "Issue",
    "Site Name", "AP Name", "MAC", "Status", "Power Constrained",
    "IP", "Gateway", "Netmask", "DNS Servers", "DHCP Server", "External IP", "Mount",
    "eth0 Up", "eth0 Speed (Mbps)", "eth0 Full Duplex",
    "eth0 TX Bytes", "eth0 TX Pkts", "eth0 RX Bytes", "eth0 RX Pkts",
    "eth0 RX Errors", "eth0 RX Peak (bps)", "eth0 TX Peak (bps)",
    "Inactive Wired VLANs",
    "Switch Name", "Switch Description", "Switch Mgmt Addr",
    "Switch Port Desc", "Switch Port ID", "Switch Chassis ID",
    "LLDP MED Supported", "PoE Req Count",
    "PoE Allocated (mW)", "PoE Requested (mW)", "PoE Draw (mW)",
    "ESL Stat",
]

def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for col, h in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def auto_width(ws):
    for col_idx, _ in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Fetching organisation info...")
    # These 3 preflight calls are accounted for in est_total_calls
    org_info  = fetch_org_info()
    org_name  = org_info.get("name", "Unknown")
    org_stats = fetch_org_stats()
    num_sites = org_stats.get("num_sites", 0)
    total_aps = (org_stats.get("num_devices_connected", 0)
                 + org_stats.get("num_devices_disconnected", 0))

    usage_used, usage_limit, usage_secs = fetch_api_usage()
    rate_remaining, secs_left = remaining_and_reset(usage_used, usage_limit, usage_secs)
    m, s = divmod(secs_left, 60)

    est_site_pages  = max(1, (num_sites + 999) // 1000)
    # 3 preflight calls already made + site pages + 1 per site
    est_run_calls   = est_site_pages + num_sites
    est_total_calls = 3 + est_run_calls

    print(f"\n{'='*60}")
    print(f"  Mist AP Details Report")
    print(f"{'='*60}")
    print(f"  Organisation:        {org_name}")
    print(f"  Org ID:              {ORG_ID}")
    print(f"  Total APs:           ~{total_aps}")
    print(f"  Total Sites:         ~{num_sites}")
    print(f"{'='*60}")
    print(f"  This report will:")
    print(f"    1. Fetch all sites         (~{est_site_pages} API call{'s' if est_site_pages != 1 else ''})")
    print(f"    2. Fetch AP stats per site (~{num_sites} API calls, one per site)")
    print(f"  Estimated total API calls:   ~{est_total_calls}")
    print(f"{'='*60}")
    print(f"  API rate limit:      {usage_limit} calls/hour")
    print(f"  Used this hour:      {usage_used}")
    print(f"  Remaining:           {rate_remaining}  (window resets in ~{m}m {s}s)")

    # --- Pre-flight rate-limit gate ---
    if rate_remaining - RATE_HEADROOM < est_run_calls:
        shortfall = est_run_calls - (rate_remaining - RATE_HEADROOM)
        print(f"\n  *** INSUFFICIENT API CALLS ***")
        print(f"  Calls still needed this run: ~{est_run_calls}")
        print(f"  Calls available (with {RATE_HEADROOM} headroom): {rate_remaining - RATE_HEADROOM}")
        print(f"  Shortfall: ~{shortfall} calls")
        print(f"  Window resets in ~{m}m {s}s")
        print(f"{'='*60}")
        print("  Options:")
        print("    w — wait for the window to reset, then start automatically")
        print("    q — quit")
        while True:
            choice = input("  Your choice (w/q): ").strip().lower()
            if choice == "q":
                print("  Aborted.")
                return
            if choice == "w":
                wait_for_reset(secs_left, reason="Waiting for rate-limit window to reset before starting.")
                usage_used, usage_limit, usage_secs = fetch_api_usage()
                rate_remaining, secs_left = remaining_and_reset(usage_used, usage_limit, usage_secs)
                print(f"  Remaining calls after reset: {rate_remaining}")
                break
            print("  Please enter 'w' to wait or 'q' to quit.")
    else:
        print(f"  Sufficient calls available.  OK to proceed.")

    print(f"{'='*60}")
    confirm = input("\n  Proceed? (y/n): ").strip().lower()
    if confirm != "y":
        print("  Aborted.")
        return

    start_time = time.time()

    # Track calls consumed during the run so we can check mid-run
    calls_this_run = 0

    print("\nFetching site list...")
    sites = fetch_all_sites()
    num_sites_actual = len(sites)
    calls_this_run += est_site_pages   # approximate; actual pages used

    print(f"\nFetching AP stats for {num_sites_actual} sites...")
    all_rows   = []
    site_times = []

    for idx, site in enumerate(sites, 1):
        site_id   = site["id"]
        site_name = site.get("name", site_id)

        # Check rate limit before each batch of ~50 sites, and always on the last one
        if idx == 1 or idx % 50 == 0:
            calls_still_needed = num_sites_actual - idx + 1
            check_rate_limit(
                calls_still_needed,
                context=f"site {idx}/{num_sites_actual}"
            )

        t0   = time.time()
        aps  = fetch_site_ap_stats(site_id, site_name, idx, num_sites_actual)
        site_times.append(time.time() - t0)
        calls_this_run += 1

        for ap in aps:
            all_rows.append(flatten_ap(ap, site_name))

    print(f"\n\n  {len(all_rows)} APs collected across {num_sites_actual} sites.")

    print("\nBuilding Excel spreadsheet...")
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Details"

    style_header(ws)

    issues_count = 0

    for r, row in enumerate(all_rows, 2):
        flags = flagged_columns(row)
        has_issue = bool(flags)
        if has_issue:
            issues_count += 1

        for c, col in enumerate(COLUMNS, 1):
            if col == "Issue":
                cell = ws.cell(row=r, column=c, value="✗" if has_issue else "✓")
                cell.fill      = FILL_ISSUE if has_issue else FILL_OK
                cell.font      = FONT_WHITE
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell = ws.cell(row=r, column=c, value=row.get(col, ""))
                if col in flags:
                    cell.fill = FILL_FLAGGED

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes   = "A2"
    auto_width(ws)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    safe_org  = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in org_name).strip().replace(" ", "_")
    filename  = f"Mist_AP_Details_{safe_org}_{timestamp}.xlsx"
    filepath  = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    elapsed  = time.time() - start_time
    avg_site = sum(site_times) / len(site_times) if site_times else 0

    print(f"\n{'='*60}")
    print(f"  Mist AP Details Summary - {org_name}")
    print(f"{'='*60}")
    print(f"  Sites processed:     {num_sites_actual}")
    print(f"  APs collected:       {len(all_rows)}")
    print(f"  APs with issues:     {issues_count}  ({len(all_rows) - issues_count} OK)")
    print(f"  Avg time/site:       {avg_site:.2f}s")
    print(f"  API calls this run:  ~{calls_this_run}")
    print(f"  Total elapsed time:  {format_eta(elapsed)}")
    print(f"{'='*60}")
    print(f"  Report saved: {filepath}")

if __name__ == "__main__":
    main()
