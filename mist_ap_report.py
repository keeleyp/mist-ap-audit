#!/usr/bin/env python3
import configparser
import requests
import time
import os
import sys
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

config = configparser.ConfigParser()
config.read(os.path.join(os.path.dirname(os.path.abspath(__file__)), "mist_ap_report.ini"))

API_BASE = config.get("mist", "api_base")
ORG_ID = config.get("mist", "org_id")
API_TOKEN = config.get("mist", "api_token")
OFFLINE_DAYS = config.getint("thresholds", "offline_days")
LOW_UPTIME_HOURS = config.getint("thresholds", "low_uptime_hours")
OUTPUT_DIR = os.path.expanduser(config.get("output", "directory"))

OFFLINE_SECONDS = OFFLINE_DAYS * 24 * 60 * 60
LOW_UPTIME_SECONDS = LOW_UPTIME_HOURS * 60 * 60
ONE_DAY_SECONDS = 24 * 60 * 60

def get_token():
    return API_TOKEN

def fetch_org_info(token):
    headers = {"Authorization": f"Token {token}"}
    resp = requests.get(f"{API_BASE}/orgs/{ORG_ID}", headers=headers)
    resp.raise_for_status()
    return resp.json()

def fetch_org_stats(token):
    headers = {"Authorization": f"Token {token}"}
    resp = requests.get(f"{API_BASE}/orgs/{ORG_ID}/stats", headers=headers)
    resp.raise_for_status()
    return resp.json()

def format_eta(seconds):
    if seconds < 60:
        return f"{int(seconds)}s"
    m, s = divmod(int(seconds), 60)
    return f"{m}m {s}s"

def fetch_all_aps(token, total_aps=None):
    headers = {"Authorization": f"Token {token}"}
    all_aps = []
    page = 1
    api_calls = 0
    page_times = []
    total_label = f"/{total_aps}" if total_aps else ""
    while True:
        url = f"{API_BASE}/orgs/{ORG_ID}/stats/devices?type=ap&limit=1000&page={page}"
        t0 = time.time()
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        page_time = time.time() - t0
        page_times.append(page_time)
        api_calls += 1
        data = resp.json()
        if not data:
            break
        all_aps.extend(data)
        avg_time = sum(page_times) / len(page_times)
        if len(data) == 1000:
            if total_aps:
                pct = min(99, len(all_aps) / total_aps * 100)
                remaining = max(0, total_aps - len(all_aps)) / 1000 * avg_time
                sys.stdout.write(f"\r  Page {page} | {len(all_aps)}{total_label} APs ({pct:.0f}%) | {avg_time:.1f}s/page | ETA: {format_eta(remaining)}   ")
            else:
                sys.stdout.write(f"\r  Page {page} | {len(all_aps)} APs | {avg_time:.1f}s/page | Elapsed: {format_eta(sum(page_times))}   ")
            sys.stdout.flush()
        else:
            sys.stdout.write(f"\r  Page {page} | {len(all_aps)}{total_label} APs | Done!                              \n")
            sys.stdout.flush()
            break
        page += 1
    total_time = sum(page_times)
    print(f"  AP fetch complete in {format_eta(total_time)}")
    return all_aps, api_calls

def fetch_site_names(token, site_ids):
    headers = {"Authorization": f"Token {token}"}
    site_map = {}
    page = 1
    api_calls = 0
    total_sites = len(site_ids)
    while True:
        url = f"{API_BASE}/orgs/{ORG_ID}/sites?limit=1000&page={page}"
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        api_calls += 1
        data = resp.json()
        if not data:
            break
        for site in data:
            site_map[site["id"]] = site.get("name", site["id"])
        sys.stdout.write(f"\r  Fetched {len(site_map)}/{total_sites} sites (page {page})   ")
        sys.stdout.flush()
        if len(data) < 1000:
            break
        page += 1
    print(f"\r  Site fetch complete - {len(site_map)} sites resolved                ")
    return site_map, api_calls

def fetch_events_24h(token):
    headers = {"Authorization": f"Token {token}"}
    now = int(time.time())
    start = now - ONE_DAY_SECONDS
    end = now
    all_events = []
    api_calls = 0
    page_times = []
    url = f"{API_BASE}/orgs/{ORG_ID}/devices/events/search?limit=1000&start={start}&end={end}&device_type=ap"
    page = 0
    total = None
    while url:
        page += 1
        t0 = time.time()
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        page_time = time.time() - t0
        page_times.append(page_time)
        api_calls += 1
        data = resp.json()
        if total is None:
            total = data.get("total", 0)
        results = data.get("results", [])
        if not results:
            break
        all_events.extend(results)
        avg_time = sum(page_times) / len(page_times)
        pct = min(99, len(all_events) / max(total, 1) * 100) if total else 0
        remaining = max(0, total - len(all_events)) / 1000 * avg_time
        sys.stdout.write(f"\r  Page {page} | {len(all_events)}/{total} events ({pct:.0f}%) | {avg_time:.1f}s/page | ETA: {format_eta(remaining)}   ")
        sys.stdout.flush()
        next_url = data.get("next")
        if next_url:
            url = f"https://api.eu.mist.com{next_url}"
        else:
            break
    print(f"\r  Events fetch complete - {len(all_events)} events in {format_eta(sum(page_times))}                    ")
    return all_events, api_calls

def count_events_by_mac(events):
    from collections import defaultdict
    counts = defaultdict(lambda: {"AP_RESTARTED": 0, "AP_DISCONNECTED": 0, "AP_CONNECTED": 0})
    for ev in events:
        mac = ev.get("mac", "")
        ev_type = ev.get("type", "")
        if ev_type in ("AP_RESTARTED", "AP_DISCONNECTED", "AP_CONNECTED"):
            counts[mac][ev_type] += 1
    return counts

def style_header(ws, headers, fill_color):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

def main():
    token = get_token()
    now = time.time()

    print("Fetching organisation info...")
    org_info = fetch_org_info(token)
    org_name = org_info.get("name", "Unknown")
    org_stats = fetch_org_stats(token)
    total_aps = org_stats.get("num_devices_connected", 0) + org_stats.get("num_devices_disconnected", 0)
    num_sites = org_stats.get("num_sites", 0)

    est_ap_pages = (total_aps // 1000) + 1
    est_site_pages = (num_sites // 1000) + 1
    est_event_pages = 50
    est_total_calls = est_ap_pages + est_site_pages + est_event_pages + 2

    print(f"\n{'='*55}")
    print(f"  Mist AP Report")
    print(f"{'='*55}")
    print(f"  Organisation:     {org_name}")
    print(f"  Org ID:           {ORG_ID}")
    print(f"  Total APs:        ~{total_aps}")
    print(f"  Total Sites:      ~{num_sites}")
    print(f"{'='*55}")
    print(f"  This report will:")
    print(f"    1. Fetch all AP stats  (~{est_ap_pages} API calls)")
    print(f"    2. Fetch site names    (~{est_site_pages} API calls)")
    print(f"    3. Fetch 24h events    (~{est_event_pages} API calls)")
    print(f"  Estimated total API calls: ~{est_total_calls}")
    print(f"{'='*55}")

    confirm = input("\n  Proceed? (y/n): ").strip().lower()
    if confirm != "y":
        print("  Aborted.")
        return

    start_time = time.time()
    print("\nFetching APs from Mist API...")
    all_aps, ap_api_calls = fetch_all_aps(token, total_aps=total_aps)
    print(f"\nTotal APs fetched: {len(all_aps)}")

    site_ids = set(ap.get("site_id", "") for ap in all_aps if ap.get("site_id"))
    print(f"Fetching names for {len(site_ids)} sites...")
    site_map, site_api_calls = fetch_site_names(token, site_ids)
    print("Site names fetched.")

    online = []
    offline_lt4 = []
    offline_gt4 = []
    low_uptime = []

    for ap in all_aps:
        status = ap.get("status", "")
        mac = ap.get("mac", "")
        row_base = {
            "Name": ap.get("name", ""),
            "Site ID": ap.get("site_id", ""),
            "Site Name": site_map.get(ap.get("site_id", ""), ""),
            "MAC": mac,
            "Serial": ap.get("serial", ""),
            "Model": ap.get("model", ""),
            "HW Rev": ap.get("hw_rev", ""),
            "Device Profile": ap.get("deviceprofile_name", ""),
            "Status": status,
        }
        uptime = ap.get("uptime", None)
        if status == "connected":
            if uptime is not None:
                row_base["Uptime (days)"] = round(uptime / 86400, 1)
            else:
                row_base["Uptime (days)"] = ""
            online.append(row_base)
            if uptime is not None and uptime < LOW_UPTIME_SECONDS:
                hours = uptime / 3600
                low_uptime.append({**row_base, "Uptime (seconds)": uptime, "Uptime (hours)": round(hours, 1)})
        else:
            last_seen = ap.get("last_seen", 0)
            days_since = (now - last_seen) / 86400 if last_seen else None
            date_str = datetime.fromtimestamp(last_seen, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC") if last_seen else ""
            row_base["EPOC Last Seen"] = last_seen
            row_base["Date Last Seen"] = date_str
            row_base["Days Since Last Seen"] = round(days_since, 1) if days_since is not None else ""
            if days_since is not None and days_since <= OFFLINE_DAYS:
                offline_lt4.append(row_base)
            else:
                offline_gt4.append(row_base)

    print(f"\nCategorised: {len(online)} online, {len(offline_lt4)} offline <4d, {len(offline_gt4)} offline >4d, {len(low_uptime)} uptime <24h")

    print("Fetching 24h of AP events...")
    events, event_api_calls = fetch_events_24h(token)
    event_counts = count_events_by_mac(events)

    for ap_list in [low_uptime, offline_lt4]:
        for ap_row in ap_list:
            mac = ap_row["MAC"]
            counts = event_counts.get(mac, {})
            ap_row["AP_RESTARTED (24h)"] = counts.get("AP_RESTARTED", 0)
            ap_row["AP_DISCONNECTED (24h)"] = counts.get("AP_DISCONNECTED", 0)
            ap_row["AP_CONNECTED (24h)"] = counts.get("AP_CONNECTED", 0)

    print("Building Excel spreadsheet...")

    wb = Workbook()

    base_headers = ["Name", "Site ID", "Site Name", "MAC", "Serial", "Model", "HW Rev", "Device Profile", "Status"]
    offline_lt4_headers = base_headers + ["EPOC Last Seen", "Date Last Seen", "Days Since Last Seen", "AP_RESTARTED (24h)", "AP_DISCONNECTED (24h)", "AP_CONNECTED (24h)"]
    offline_gt4_headers = base_headers + ["EPOC Last Seen", "Date Last Seen", "Days Since Last Seen"]

    online_headers = base_headers + ["Uptime (days)"]

    ws1 = wb.active
    ws1.title = "Online APs"
    style_header(ws1, online_headers, "2E7D32")
    for r, ap in enumerate(online, 2):
        for c, h in enumerate(online_headers, 1):
            ws1.cell(row=r, column=c, value=ap.get(h, ""))
    auto_width(ws1)

    ws2 = wb.create_sheet(f"Offline < {OFFLINE_DAYS} Days")
    style_header(ws2, offline_lt4_headers, "E65100")
    for r, ap in enumerate(offline_lt4, 2):
        for c, h in enumerate(offline_lt4_headers, 1):
            ws2.cell(row=r, column=c, value=ap.get(h, ""))
    auto_width(ws2)

    ws3 = wb.create_sheet(f"Offline > {OFFLINE_DAYS} Days")
    style_header(ws3, offline_gt4_headers, "B71C1C")
    for r, ap in enumerate(offline_gt4, 2):
        for c, h in enumerate(offline_gt4_headers, 1):
            ws3.cell(row=r, column=c, value=ap.get(h, ""))
    auto_width(ws3)

    uptime_headers = base_headers + ["Uptime (seconds)", "Uptime (hours)", "AP_RESTARTED (24h)", "AP_DISCONNECTED (24h)", "AP_CONNECTED (24h)"]
    ws4 = wb.create_sheet(f"Uptime < {LOW_UPTIME_HOURS} Hours")
    style_header(ws4, uptime_headers, "1565C0")
    for r, ap in enumerate(low_uptime, 2):
        for c, h in enumerate(uptime_headers, 1):
            ws4.cell(row=r, column=c, value=ap.get(h, ""))
    auto_width(ws4)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    safe_org_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in org_name).strip().replace(" ", "_")
    filename = f"Mist_AP_Report_{safe_org_name}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    total_api = ap_api_calls + site_api_calls + event_api_calls
    print(f"\n{'='*55}")
    print(f"  Mist AP Report Summary - {org_name}")
    print(f"{'='*55}")
    print(f"  Total APs processed:        {len(all_aps)}")
    print(f"  Online APs (Sheet 1):       {len(online)}")
    print(f"  Offline < {OFFLINE_DAYS} days (Sheet 2): {len(offline_lt4)}")
    print(f"  Offline > {OFFLINE_DAYS} days (Sheet 3): {len(offline_gt4)}")
    print(f"  Uptime < {LOW_UPTIME_HOURS}h (Sheet 4):     {len(low_uptime)}")
    print(f"  Unique sites:               {len(site_ids)}")
    print(f"  Events fetched (24h):       {len(events)}")
    print(f"{'='*50}")
    print(f"  API calls - AP pages:       {ap_api_calls}")
    print(f"  API calls - Site pages:     {site_api_calls}")
    print(f"  API calls - Event pages:    {event_api_calls}")
    print(f"  Total API calls:            {total_api}")
    elapsed = time.time() - start_time
    print(f"  Total elapsed time:       {format_eta(elapsed)}")
    print(f"{'='*55}")
    print(f"  Report saved: {filepath}")

if __name__ == "__main__":
    main()
